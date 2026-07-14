#!/usr/bin/env python3
"""Create or update the part definitions Excel file (part definitions.xlsx)."""

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

PROJECT_DIR = Path(__file__).parent
PART_DEFINITIONS_XLSX = PROJECT_DIR / "part definitions.xlsx"

PART_CATEGORIES = [
    "MECH",
    "ELECT",
    "3DPRT",
    "CNC",
    "PCB",
]

MODULE_TYPES = [
    "Tip Holder 1000",
    "Tip Holder 100",
    "MO:HEAT",
    "Tube Holder 8T2",
    "Tube Holder 888",
    "MO:TILT PASSIVE",
    "MO:TILT",
    "MO:COOL 2ml",
    "MO:COOL 15m",
    "MO:CROSCOPE",
    "MO:SHAKE",
]

# Part list and categories (matches part definitions.xlsx)
PARTS = [
    ("Bottom Plate", "CNC"),
    ("Top Plate", "CNC"),
    ("Cooling Tubes Block", "CNC"),
    ("50 ml Holder", "CNC"),
    ("8 ml Holder", "CNC"),
    ("Trash", "3DPRT"),
    ("Magnets", "MECH"),
    ("Spring Plungers 6 mm", "MECH"),
    ("Spring Plungers 5 mm", "MECH"),
    ("Thermal Mat Adhesive", "MECH"),
    ("Thermal Mat 1mm", "MECH"),
    ("Thermal Paste", "MECH"),
    ("Mill-Max Connector", "ELECT"),
    ("Temperature Sensor", "ELECT"),
    ("Heating Mat", "ELECT"),
    ("Fan Big", "ELECT"),
    ("Fan Small", "ELECT"),
    ("Peltier", "ELECT"),
    ("Temperature Fuse", "ELECT"),
    ("Voltage Step Down", "ELECT"),
    ("Frog Passive", "PCB"),
    ("Frog Active", "PCB"),
    ("Heater Body", "3DPRT"),
    ("Tilting Body", "3DPRT"),
    ("Tip Holder Bottom", "3DPRT"),
    ("Tip Holder Top 1000", "3DPRT"),
    ("Tip Holder Top 100", "3DPRT"),
    ("Tube Holder Bottom", "3DPRT"),
    ("Tube Holder Top", "3DPRT"),
    ("Tube Holder Trash", "3DPRT"),
    ("Cooler Bottom", "3DPRT"),
    ("Cooler Top", "3DPRT"),
    ("Cooler Screw Cover", "3DPRT"),
    ("Frog Holder", "3DPRT"),
    ("Magnet Stopper", "3DPRT"),
]

# Parts every module type needs (qty per module)
COMMON_MODULE_PARTS = {
    "Frog Holder": 1,
    "Bottom Plate": 1,
    "Magnets": 4,
    "Spring Plungers 6 mm": 4,
    "Magnet Stopper": 4,
}

# Each module gets one PCB type (not both)
MODULE_FROG_PCB = {
    "Tip Holder 1000": "Frog Passive",
    "Tip Holder 100": "Frog Passive",
    "MO:HEAT": "Frog Active",
    "Tube Holder 8T2": "Frog Passive",
    "Tube Holder 888": "Frog Passive",
    "MO:TILT PASSIVE": "Frog Passive",
    "MO:TILT": "Frog Active",
    "MO:COOL 2ml": "Frog Active",
    "MO:COOL 15m": "Frog Active",
    "MO:CROSCOPE": "Frog Active",
    "MO:SHAKE": "Frog Active",
}

# Module-specific parts (common parts above apply to all modules too)
BOM_DEFAULTS = [
    ("Top Plate", "MO:HEAT", 1),
    ("Top Plate", "MO:TILT PASSIVE", 1),
    ("Cooling Tubes Block", "MO:COOL 2ml", 1),
    ("50 ml Holder", "Tube Holder 8T2", 1),
    ("8 ml Holder", "Tube Holder 8T2", 1),
    ("8 ml Holder", "Tube Holder 888", 3),
    ("Trash", "Tube Holder 8T2", 1),
    ("Trash", "Tube Holder 888", 1),
    ("Spring Plungers 5 mm", "MO:HEAT", 3),
    ("Thermal Mat Adhesive", "MO:HEAT", 1),
    ("Thermal Mat 1mm", "MO:COOL 2ml", 1),
    ("Thermal Paste", "MO:COOL 2ml", 1),
    ("Mill-Max Connector", "Tip Holder 1000", 1),
    ("Mill-Max Connector", "Tip Holder 100", 1),
    ("Mill-Max Connector", "MO:HEAT", 1),
    ("Mill-Max Connector", "Tube Holder 8T2", 1),
    ("Mill-Max Connector", "Tube Holder 888", 1),
    ("Mill-Max Connector", "MO:TILT PASSIVE", 1),
    ("Mill-Max Connector", "MO:COOL 2ml", 1),
    ("Temperature Sensor", "MO:HEAT", 1),
    ("Temperature Sensor", "MO:COOL 2ml", 1),
    ("Heating Mat", "MO:HEAT", 1),
    ("Fan Big", "MO:COOL 2ml", 1),
    ("Fan Small", "MO:COOL 2ml", 1),
    ("Peltier", "MO:COOL 2ml", 1),
    ("Temperature Fuse", "MO:HEAT", 1),
    ("Voltage Step Down", "MO:COOL 2ml", 1),
    ("Heater Body", "MO:HEAT", 1),
    ("Tilting Body", "MO:TILT PASSIVE", 1),
    ("Tip Holder Bottom", "Tip Holder 1000", 1),
    ("Tip Holder Bottom", "Tip Holder 100", 1),
    ("Tip Holder Top 1000", "Tip Holder 1000", 1),
    ("Tip Holder Top 100", "Tip Holder 100", 1),
    ("Tube Holder Bottom", "Tube Holder 8T2", 1),
    ("Tube Holder Bottom", "Tube Holder 888", 1),
    ("Tube Holder Top", "Tube Holder 8T2", 1),
    ("Tube Holder Top", "Tube Holder 888", 1),
    ("Tube Holder Trash", "Tube Holder 8T2", 1),
    ("Tube Holder Trash", "Tube Holder 888", 1),
    ("Cooler Bottom", "MO:COOL 2ml", 1),
    ("Cooler Top", "MO:COOL 2ml", 1),
    ("Cooler Screw Cover", "MO:COOL 2ml", 2),
]


def _default_values() -> dict[tuple[str, str], int]:
    values: dict[tuple[str, str], int] = {}
    for module in MODULE_TYPES:
        for part, qty in COMMON_MODULE_PARTS.items():
            values[(part, module)] = qty
        pcb = MODULE_FROG_PCB[module]
        values[(pcb, module)] = 1
    for part, module, qty in BOM_DEFAULTS:
        values[(part, module)] = qty
    return values


def _apply_pcb_rules(values: dict[tuple[str, str], int]) -> None:
    """Ensure each module has exactly one frog PCB type."""
    for module, pcb in MODULE_FROG_PCB.items():
        values[(pcb, module)] = 1
        other = "Frog Active" if pcb == "Frog Passive" else "Frog Passive"
        values.pop((other, module), None)


def _load_existing_values() -> tuple[dict[tuple[str, str], int], dict[str, str]]:
    """Load quantities and categories from existing part definitions.xlsx if present."""
    if not PART_DEFINITIONS_XLSX.exists():
        return {}, {}
    return _load_existing_from_xlsx()


def _load_existing_from_xlsx() -> tuple[dict[tuple[str, str], int], dict[str, str]]:
    """Load quantities and categories from existing part definitions.xlsx if present."""
    values: dict[tuple[str, str], int] = {}
    categories: dict[str, str] = {}

    wb = load_workbook(PART_DEFINITIONS_XLSX, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return values, categories

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    module_cols = [h for h in headers[2:] if h]

    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        part = str(row[0]).strip()
        if not part:
            continue
        if row[1] is not None and str(row[1]).strip():
            categories[part] = str(row[1]).strip()
        for module, cell in zip(module_cols, row[2:]):
            if cell is None or cell == "":
                continue
            values[(part, module)] = int(cell)

    return values, categories


def _resolve_parts(categories: dict[str, str]) -> list[tuple[str, str]]:
    """Build part list, keeping Excel order when an existing file is loaded."""
    parts_by_name = dict(PARTS)
    if PART_DEFINITIONS_XLSX.exists():
        wb = load_workbook(PART_DEFINITIONS_XLSX, data_only=True)
        ws = wb.active
        parts = []
        seen = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            part = str(row[0]).strip()
            if not part or part in seen:
                continue
            seen.add(part)
            category = categories.get(part) or (str(row[1]).strip() if row[1] else "")
            if not category:
                category = parts_by_name.get(part, "MECH")
            parts.append((part, category))
        for part, category in PARTS:
            if part not in seen:
                parts.append((part, category))
        if parts:
            return parts
    return PARTS


def create_part_definitions() -> Path:
    defaults = _default_values()
    categories: dict[str, str] = {}

    existing, categories = _load_existing_values()

    values = {**defaults, **existing}
    _apply_pcb_rules(values)
    parts = _resolve_parts(categories)

    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    headers = ["Part", "Category", *MODULE_TYPES]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for part, category in parts:
        row = [part, category]
        for module in MODULE_TYPES:
            row.append(values.get((part, module), ""))
        ws.append(row)

    last_row = ws.max_row
    last_col = ws.max_column
    table_ref = f"A1:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName="BOMTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    category_list = ",".join(PART_CATEGORIES)
    category_validation = DataValidation(
        type="list",
        formula1=f'"{category_list}"',
        allow_blank=False,
    )
    category_validation.error = "Choose a category from the list"
    category_validation.errorTitle = "Invalid category"
    ws.add_data_validation(category_validation)
    category_validation.add(f"B2:B{last_row}")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    for col in range(3, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    ws.freeze_panes = "C2"

    for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=3, max_col=last_col):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    wb.save(PART_DEFINITIONS_XLSX)
    return PART_DEFINITIONS_XLSX


if __name__ == "__main__":
    path = create_part_definitions()
    print(f"Created: {path}")
