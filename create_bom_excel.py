#!/usr/bin/env python3
"""Create or update the part definitions Excel file (part definitions.xlsx)."""

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

PROJECT_DIR = Path(__file__).parent
PART_DEFINITIONS_XLSX = PROJECT_DIR / "part definitions.xlsx"

PART_CATEGORIES = [
    "mechanical part",
    "electronics",
    "3d printing",
    "cnc metal",
    "pcb",
]

MODULE_TYPES = [
    "tip holder 1000",
    "tip holder 100",
    "MO:HEAT",
    "tube holder 8T2",
    "tube holder 888",
    "MO:TILT PASSIVE",
    "MO:TILT",
    "MO:COOL 2ml",
    "MO:COOL 15m",
    "MO:CROSCOPE",
    "MO:SHAKE",
]

# Part list and categories (matches part definitions.xlsx)
PARTS = [
    ("bottom plate", "cnc metal"),
    ("top plate", "cnc metal"),
    ("cooling tubes block", "cnc metal"),
    ("50 ml holder", "cnc metal"),
    ("8 ml holder", "cnc metal"),
    ("trash", "3d printing"),
    ("Magnets", "mechanical part"),
    ("Spring plungers 6 mm", "mechanical part"),
    ("Spring plungers 5 mm", "mechanical part"),
    ("Thermal Mat adhesive", "mechanical part"),
    ("Thermal Mat 1mm", "mechanical part"),
    ("Thermal Paste", "mechanical part"),
    ("mill-max connector", "electronics"),
    ("temperature sensor", "electronics"),
    ("heating mat", "electronics"),
    ("fan big", "electronics"),
    ("fan small", "electronics"),
    ("Peltier", "electronics"),
    ("Temperature Fuse", "electronics"),
    ("Voltage step down", "electronics"),
    ("frog passive", "pcb"),
    ("frog active", "pcb"),
    ("heater body", "3d printing"),
    ("tilting body", "3d printing"),
    ("tip holder bottom", "3d printing"),
    ("tip holder top 1000", "3d printing"),
    ("tip holder top 100", "3d printing"),
    ("tube holder bottom", "3d printing"),
    ("tube holder top", "3d printing"),
    ("Tube holder Trash", "3d printing"),
    ("cooler bottom", "3d printing"),
    ("cooler top", "3d printing"),
    ("Cooler screw cover", "3d printing"),
    ("frog holder", "3d printing"),
    ("magnet stopper", "3d printing"),
]

# Parts every module type needs (qty per module)
COMMON_MODULE_PARTS = {
    "frog holder": 1,
    "bottom plate": 1,
    "Magnets": 4,
    "Spring plungers 6 mm": 4,
    "magnet stopper": 4,
}

# Each module gets one PCB type (not both)
MODULE_FROG_PCB = {
    "tip holder 1000": "frog passive",
    "tip holder 100": "frog passive",
    "MO:HEAT": "frog active",
    "tube holder 8T2": "frog passive",
    "tube holder 888": "frog passive",
    "MO:TILT PASSIVE": "frog passive",
    "MO:TILT": "frog active",
    "MO:COOL 2ml": "frog active",
    "MO:COOL 15m": "frog active",
    "MO:CROSCOPE": "frog active",
    "MO:SHAKE": "frog active",
}

# Module-specific parts (common parts above apply to all modules too)
BOM_DEFAULTS = [
    ("top plate", "MO:HEAT", 1),
    ("top plate", "MO:TILT PASSIVE", 1),
    ("cooling tubes block", "MO:COOL 2ml", 1),
    ("50 ml holder", "tube holder 8T2", 1),
    ("8 ml holder", "tube holder 8T2", 1),
    ("8 ml holder", "tube holder 888", 3),
    ("trash", "tube holder 8T2", 1),
    ("trash", "tube holder 888", 1),
    ("Spring plungers 5 mm", "MO:HEAT", 3),
    ("Thermal Mat adhesive", "MO:HEAT", 1),
    ("Thermal Mat 1mm", "MO:COOL 2ml", 1),
    ("Thermal Paste", "MO:COOL 2ml", 1),
    ("mill-max connector", "tip holder 1000", 1),
    ("mill-max connector", "tip holder 100", 1),
    ("mill-max connector", "MO:HEAT", 1),
    ("mill-max connector", "tube holder 8T2", 1),
    ("mill-max connector", "tube holder 888", 1),
    ("mill-max connector", "MO:TILT PASSIVE", 1),
    ("mill-max connector", "MO:COOL 2ml", 1),
    ("temperature sensor", "MO:HEAT", 1),
    ("temperature sensor", "MO:COOL 2ml", 1),
    ("heating mat", "MO:HEAT", 1),
    ("fan big", "MO:COOL 2ml", 1),
    ("fan small", "MO:COOL 2ml", 1),
    ("Peltier", "MO:COOL 2ml", 1),
    ("Temperature Fuse", "MO:HEAT", 1),
    ("Voltage step down", "MO:COOL 2ml", 1),
    ("heater body", "MO:HEAT", 1),
    ("tilting body", "MO:TILT PASSIVE", 1),
    ("tip holder bottom", "tip holder 1000", 1),
    ("tip holder bottom", "tip holder 100", 1),
    ("tip holder top 1000", "tip holder 1000", 1),
    ("tip holder top 100", "tip holder 100", 1),
    ("tube holder bottom", "tube holder 8T2", 1),
    ("tube holder bottom", "tube holder 888", 1),
    ("tube holder top", "tube holder 8T2", 1),
    ("tube holder top", "tube holder 888", 1),
    ("Tube holder Trash", "tube holder 8T2", 1),
    ("Tube holder Trash", "tube holder 888", 1),
    ("cooler bottom", "MO:COOL 2ml", 1),
    ("cooler top", "MO:COOL 2ml", 1),
    ("Cooler screw cover", "MO:COOL 2ml", 2),
]


def _default_values() -> dict[tuple[str, str], int]:
    values: dict[tuple[str, str], int] = {}
    for module in MODULE_TYPES:
        for part, qty in COMMON_MODULE_PARTS.items():
            values[(part, module)] = qty
        pcb = MODULE_FROG_PCB[module]
        values[(pcb, module)] = 1
    for part, module, qty in BOM_DEFAULTS:
        values[(part, module)] = qty
    return values


def _apply_pcb_rules(values: dict[tuple[str, str], int]) -> None:
    """Ensure each module has exactly one frog PCB type."""
    for module, pcb in MODULE_FROG_PCB.items():
        values[(pcb, module)] = 1
        other = "frog active" if pcb == "frog passive" else "frog passive"
        values.pop((other, module), None)


def _load_existing_values() -> tuple[dict[tuple[str, str], int], dict[str, str]]:
    """Load quantities and categories from existing part definitions.xlsx if present."""
    if not PART_DEFINITIONS_XLSX.exists():
        return {}, {}
    return _load_existing_from_xlsx()


def _load_existing_from_xlsx() -> tuple[dict[tuple[str, str], int], dict[str, str]]:
    """Load quantities and categories from existing part definitions.xlsx if present."""
    values: dict[tuple[str, str], int] = {}
    categories: dict[str, str] = {}

    wb = load_workbook(PART_DEFINITIONS_XLSX, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return values, categories

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    module_cols = [h for h in headers[2:] if h]

    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        part = str(row[0]).strip()
        if not part:
            continue
        if row[1] is not None and str(row[1]).strip():
            categories[part] = str(row[1]).strip()
        for module, cell in zip(module_cols, row[2:]):
            if cell is None or cell == "":
                continue
            values[(part, module)] = int(cell)

    return values, categories


def _resolve_parts(categories: dict[str, str]) -> list[tuple[str, str]]:
    """Build part list, keeping Excel order when an existing file is loaded."""
    parts_by_name = dict(PARTS)
    if PART_DEFINITIONS_XLSX.exists():
        wb = load_workbook(PART_DEFINITIONS_XLSX, data_only=True)
        ws = wb.active
        parts = []
        seen = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            part = str(row[0]).strip()
            if not part or part in seen:
                continue
            seen.add(part)
            category = categories.get(part) or (str(row[1]).strip() if row[1] else "")
            if not category:
                category = parts_by_name.get(part, "mechanical part")
            parts.append((part, category))
        for part, category in PARTS:
            if part not in seen:
                parts.append((part, category))
        if parts:
            return parts
    return PARTS


def create_bom_excel() -> Path:
    defaults = _default_values()
    categories: dict[str, str] = {}

    existing, categories = _load_existing_values()

    values = {**defaults, **existing}
    _apply_pcb_rules(values)
    parts = _resolve_parts(categories)

    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    headers = ["Part", "Category", *MODULE_TYPES]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for part, category in parts:
        row = [part, category]
        for module in MODULE_TYPES:
            row.append(values.get((part, module), ""))
        ws.append(row)

    last_row = ws.max_row
    last_col = ws.max_column
    table_ref = f"A1:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName="BOMTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    category_list = ",".join(PART_CATEGORIES)
    category_validation = DataValidation(
        type="list",
        formula1=f'"{category_list}"',
        allow_blank=False,
    )
    category_validation.error = "Choose a category from the list"
    category_validation.errorTitle = "Invalid category"
    ws.add_data_validation(category_validation)
    category_validation.add(f"B2:B{last_row}")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    for col in range(3, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    ws.freeze_panes = "C2"

    for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=3, max_col=last_col):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    wb.save(PART_DEFINITIONS_XLSX)
    return PART_DEFINITIONS_XLSX


if __name__ == "__main__":
    path = create_bom_excel()
    print(f"Created: {path}")
