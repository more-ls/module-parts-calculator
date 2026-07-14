#!/usr/bin/env python3
"""Create or update the inventory Excel file (inventory.xlsx)."""

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from create_part_definitions import PARTS, PART_CATEGORIES, PART_DEFINITIONS_XLSX

PROJECT_DIR = Path(__file__).parent
INVENTORY_XLSX = PROJECT_DIR / "inventory.xlsx"


def _load_parts_from_definitions() -> list[tuple[str, str]]:
    """Load part list from part definitions.xlsx, falling back to code defaults."""
    if not PART_DEFINITIONS_XLSX.exists():
        return PARTS

    wb = load_workbook(PART_DEFINITIONS_XLSX, data_only=True)
    ws = wb.active
    parts = []
    seen = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        part = str(row[0]).strip()
        if not part or part in seen:
            continue
        seen.add(part)
        category = str(row[1]).strip() if row[1] is not None else ""
        parts.append((part, category or dict(PARTS).get(part, "MECH")))

    for part, category in PARTS:
        if part not in seen:
            parts.append((part, category))

    return parts


def _load_existing_stock() -> dict[str, int]:
    """Load current stock counts from existing inventory.xlsx if present."""
    stock: dict[str, int] = {}
    if not INVENTORY_XLSX.exists():
        return stock

    wb = load_workbook(INVENTORY_XLSX, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return stock

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    try:
        part_idx = headers.index("Part")
        stock_idx = headers.index("Stock Count")
    except ValueError:
        return stock

    for row in rows[1:]:
        if not row or row[part_idx] is None:
            continue
        part = str(row[part_idx]).strip()
        if not part:
            continue
        value = row[stock_idx] if stock_idx < len(row) else None
        if value is None or value == "":
            continue
        stock[part] = int(value)

    return stock


def create_inventory_excel() -> Path:
    parts = _load_parts_from_definitions()
    existing_stock = _load_existing_stock()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    headers = ["Part", "Category", "Stock Count"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for part, category in parts:
        ws.append([part, category, existing_stock.get(part, 0)])

    last_row = ws.max_row
    table_ref = f"A1:{get_column_letter(len(headers))}{last_row}"
    table = Table(displayName="InventoryTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    category_list = ",".join(PART_CATEGORIES)
    category_validation = DataValidation(
        type="list",
        formula1=f'"{category_list}"',
        allow_blank=False,
    )
    category_validation.error = "Choose a category from the list"
    category_validation.errorTitle = "Invalid category"
    ws.add_data_validation(category_validation)
    category_validation.add(f"B2:B{last_row}")

    stock_validation = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=True,
    )
    stock_validation.error = "Stock count must be zero or a positive whole number"
    stock_validation.errorTitle = "Invalid stock count"
    ws.add_data_validation(stock_validation)
    stock_validation.add(f"C2:C{last_row}")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.freeze_panes = "A2"

    for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    wb.save(INVENTORY_XLSX)
    return INVENTORY_XLSX


if __name__ == "__main__":
    path = create_inventory_excel()
    print(f"Created: {path}")
