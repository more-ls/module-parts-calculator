#!/usr/bin/env python3
"""Calculate total modules and parts needed from module type quantities."""

import os
import re
import subprocess
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from fpdf import FPDF
from fpdf.enums import XPos, YPos
from openpyxl import load_workbook

from create_part_definitions import PART_CATEGORIES

PROJECT_DIR = Path(__file__).parent
PART_DEFINITIONS_XLSX = PROJECT_DIR / "part definitions.xlsx"
INVENTORY_XLSX = PROJECT_DIR / "inventory.xlsx"
REPORTS_DIR = PROJECT_DIR / "reports"


PARTS_TABLE_HEADERS = [
    "Part",
    "Type",
    "Needed",
    "In Stock",
    "Remaining",
    "Status",
]

# Fixed widths from auto-sizing; Remaining and Status swapped.
PARTS_TABLE_COL_WIDTHS = (49.13, 20.20, 23.58, 24.92, 20.89, 51.28)

MODULES_TABLE_HEADERS = ["Module Type", "Count"]
CATEGORY_SUMMARY_TABLE_HEADERS = ["Part Type", "Needed", "In Stock", "To Order"]
MODULES_STYLE_TABLE_HEADERS = (MODULES_TABLE_HEADERS, CATEGORY_SUMMARY_TABLE_HEADERS)


@dataclass
class ReportSection:
    title: str
    headers: list[str] = field(default_factory=list)
    rows: list[list[str]] = field(default_factory=list)
    intro: list[str] = field(default_factory=list)
    intro_table_headers: list[str] = field(default_factory=list)
    intro_table_rows: list[list[str]] = field(default_factory=list)
    footer_table_headers: list[str] = field(default_factory=list)
    footer_table_rows: list[list[str]] = field(default_factory=list)
    footer: list[str] = field(default_factory=list)
    group: str = ""
    page_break_before: bool = False


def _parse_qty(raw) -> int | None:
    if raw is None or raw == "":
        return None
    raw = str(raw).strip()
    if not raw:
        return None
    try:
        return int(float(raw))
    except ValueError as exc:
        raise ValueError(f"Invalid quantity '{raw}'") from exc


def load_bom_from_excel(
    xlsx_path: Path = PART_DEFINITIONS_XLSX,
) -> tuple[list[str], dict[str, str], dict[str, dict[str, int]]]:
    """
    Load the BOM matrix directly from the Excel part definitions file.

    Returns:
        module_types: ordered list of module type names (column headers)
        part_category: {part_name: category}
        module_types_bom: {module_type: {part_name: qty_per_module}}
    """
    if not xlsx_path.exists():
        from create_part_definitions import create_part_definitions

        create_part_definitions()

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Excel file is empty: {xlsx_path}")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    if headers[0] != "Part" or headers[1] != "Category":
        raise ValueError("Part definitions must start with columns: Part, Category, then module types")

    module_types = [h for h in headers[2:] if h]
    part_category: dict[str, str] = {}
    module_types_bom: dict[str, dict[str, int]] = {m: {} for m in module_types}

    for row in rows[1:]:
        if not row or row[0] is None or str(row[0]).strip() == "":
            continue

        part = str(row[0]).strip()
        part_category[part] = str(row[1]).strip() if row[1] is not None else ""

        for module_type, cell in zip(module_types, row[2:]):
            try:
                qty = _parse_qty(cell)
            except ValueError as exc:
                raise ValueError(
                    f"Invalid quantity for part '{part}' in module '{module_type}': {exc}"
                ) from exc
            if qty is not None and qty > 0:
                module_types_bom[module_type][part] = qty

    return module_types, part_category, module_types_bom


def load_inventory_from_excel(
    xlsx_path: Path = INVENTORY_XLSX,
) -> dict[str, int]:
    """Load current stock counts from inventory.xlsx."""
    if not xlsx_path.exists():
        from create_inventory_excel import create_inventory_excel

        create_inventory_excel()

    stock: dict[str, int] = {}
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return stock

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    try:
        part_idx = headers.index("Part")
        stock_idx = headers.index("Stock Count")
    except ValueError:
        return stock

    for row in rows[1:]:
        if not row or row[part_idx] is None:
            continue
        part = str(row[part_idx]).strip()
        if not part:
            continue
        value = row[stock_idx] if stock_idx < len(row) else None
        if value is None or value == "":
            stock[part] = 0
        else:
            stock[part] = int(float(value))

    return stock


def _remaining_to_order(needed: int, stock: int) -> int:
    return max(needed - stock, 0)


def get_project_name() -> str:
    """Prompt the user for a company or project name."""
    while True:
        name = input("Enter company / project name: ").strip()
        if name:
            return name
        print("  Please enter a name.")


def get_module_counts(module_type_names: list[str]) -> dict[str, int]:
    """Prompt the user for how many of each module type to build."""
    counts: dict[str, int] = {}

    print("Enter the number of modules for each type (0 to skip):\n")
    for module_type in module_type_names:
        while True:
            raw = input(f"  {module_type}: ").strip()
            if raw == "":
                counts[module_type] = 0
                break
            try:
                value = int(raw)
                if value < 0:
                    print("    Please enter a non-negative number.")
                    continue
                counts[module_type] = value
                break
            except ValueError:
                print("    Please enter a valid integer.")

    return counts


def calculate(
    module_counts: dict[str, int],
    module_types_bom: dict[str, dict[str, int]],
) -> tuple[int, dict[str, int], dict[str, dict[str, int]]]:
    """Compute totals from module counts."""
    total_modules = sum(module_counts.values())
    total_parts: dict[str, int] = defaultdict(int)
    parts_by_type: dict[str, dict[str, int]] = {}

    for module_type, count in module_counts.items():
        if count == 0:
            parts_by_type[module_type] = {}
            continue

        type_parts: dict[str, int] = {}
        for part, qty_per_module in module_types_bom[module_type].items():
            needed = count * qty_per_module
            type_parts[part] = needed
            total_parts[part] += needed

        parts_by_type[module_type] = type_parts

    return total_modules, dict(total_parts), parts_by_type


def _category_sort_key(category: str) -> int:
    try:
        return PART_CATEGORIES.index(category)
    except ValueError:
        return len(PART_CATEGORIES)


def _format_display_name(name: str) -> str:
    """Format part and category names consistently for reports."""
    text = str(name).strip().title()
    text = re.sub(r"(\d+)Mm\b", r"\1mm", text)
    text = re.sub(r"(\d+)Ml\b", r"\1ml", text)
    text = text.replace(" Mm", " mm").replace(" Ml", " ml")
    text = text.replace("Cnc Metal", "CNC Metal").replace("Cnc ", "CNC ")
    return text.replace("Pcb", "PCB")


def _format_module_type(name: str) -> str:
    """Format module type names, preserving MO: prefixed codes as-is."""
    text = str(name).strip()
    if text.upper().startswith("MO:"):
        return text
    return _format_display_name(text)


def _parts_sorted_by_category(
    parts: dict[str, int],
    part_category: dict[str, str],
) -> list[tuple[str, str, int]]:
    """Return (category, part, qty) tuples sorted by category type then part name."""
    rows = [
        (part_category.get(part, "other"), part, qty)
        for part, qty in parts.items()
    ]
    return sorted(rows, key=lambda item: (_category_sort_key(item[0]), item[1]))


def build_summary_report(
    project_name: str,
    module_type_names: list[str],
    part_category: dict[str, str],
    module_counts: dict[str, int],
    total_modules: int,
    total_parts: dict[str, int],
    inventory_stock: dict[str, int],
) -> tuple[str, list[ReportSection]]:
    """Build structured report sections for console and PDF output."""
    title = f"{project_name} - MODULE & PARTS SUMMARY"
    sections: list[ReportSection] = []
    grand_total = sum(total_parts.values())
    grand_in_stock = sum(inventory_stock.get(part, 0) for part in total_parts)
    grand_to_order = sum(
        _remaining_to_order(qty, inventory_stock.get(part, 0))
        for part, qty in total_parts.items()
    )

    total_rows = []
    for category, part, qty in _parts_sorted_by_category(total_parts, part_category):
        in_stock = inventory_stock.get(part, 0)
        to_order = _remaining_to_order(qty, in_stock)
        total_rows.append([
            _format_display_name(part),
            category,
            str(qty),
            str(in_stock),
            str(to_order),
            "",
        ])

    category_summary_rows = []
    for category in PART_CATEGORIES:
        needed_subtotal = sum(
            qty for part, qty in total_parts.items() if part_category.get(part) == category
        )
        if needed_subtotal:
            in_stock_subtotal = sum(
                inventory_stock.get(part, 0)
                for part, qty in total_parts.items()
                if part_category.get(part) == category
            )
            to_order_subtotal = sum(
                _remaining_to_order(qty, inventory_stock.get(part, 0))
                for part, qty in total_parts.items()
                if part_category.get(part) == category
            )
            category_summary_rows.append([
                category,
                str(needed_subtotal),
                str(in_stock_subtotal),
                str(to_order_subtotal),
            ])
    category_summary_rows.append([
        "GRAND TOTAL",
        str(grand_total),
        str(grand_in_stock),
        str(grand_to_order),
    ])

    module_rows = [
        [_format_module_type(module_type), str(module_counts.get(module_type, 0))]
        for module_type in module_type_names
        if module_counts.get(module_type, 0) > 0
    ]
    module_rows.append(["Total", str(total_modules)])
    sections.append(
        ReportSection(
            title="Modules Summary",
            group="Modules Summary",
            headers=MODULES_TABLE_HEADERS,
            rows=module_rows,
            footer_table_headers=CATEGORY_SUMMARY_TABLE_HEADERS,
            footer_table_rows=category_summary_rows,
        )
    )

    sections.append(
        ReportSection(
            title=f"Total Parts (needed {grand_total})",
            group="Parts Summary",
            page_break_before=True,
            headers=PARTS_TABLE_HEADERS,
            rows=total_rows,
        )
    )

    order_rows = [row for row in total_rows if int(row[4]) > 0]
    if order_rows:
        sections.append(
            ReportSection(
                title="Parts to Order",
                group="Parts to Order",
                page_break_before=True,
                headers=PARTS_TABLE_HEADERS,
                rows=order_rows,
            )
        )

    return title, sections


def _safe_filename(name: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*]+', "", name).strip()
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned or "report"


def _report_pdf_path(project_name: str) -> Path:
    """Pick a PDF path, using a new suffix if today's report file is open elsewhere."""
    REPORTS_DIR.mkdir(exist_ok=True)
    report_date = date.today().isoformat()
    stem = f"{_safe_filename(project_name)}_{report_date}"

    for n in range(1, 100):
        suffix = "" if n == 1 else f"_{n}"
        path = REPORTS_DIR / f"{stem}{suffix}.pdf"
        if not path.exists():
            return path
        try:
            with path.open("r+b"):
                return path
        except PermissionError:
            continue

    stamp = datetime.now().strftime("%H%M%S")
    return REPORTS_DIR / f"{stem}_{stamp}.pdf"


def _table_column_alignments(headers: list[str]) -> tuple[tuple[str, ...], set[int]]:
    """Return column alignments and text columns that keep full values."""
    if headers == PARTS_TABLE_HEADERS:
        return (
            ("JUSTIFY", "JUSTIFY", "CENTER", "CENTER", "CENTER", "CENTER"),
            {0, 1},
        )
    if headers == MODULES_TABLE_HEADERS:
        return (
            ("JUSTIFY", "CENTER"),
            {0, 1},
        )
    if headers == CATEGORY_SUMMARY_TABLE_HEADERS:
        return (
            ("JUSTIFY", "CENTER", "CENTER", "CENTER"),
            {0},
        )
    return tuple("CENTER" for _ in headers), set()


def _auto_col_widths(
    pdf: FPDF,
    headers: list[str],
    rows: list[list[str]],
    *,
    table_width: float,
    cell_padding: float = 4.0,
    shrink_to_fit: bool = False,
) -> tuple[float, ...]:
    """Size columns from header and cell text, then scale to the table width."""
    max_widths = [0.0] * len(headers)

    pdf.set_font("Helvetica", "B", 9)
    for col_idx, header in enumerate(headers):
        max_widths[col_idx] = max(
            max_widths[col_idx],
            pdf.get_string_width(header) + cell_padding,
        )

    for row in rows:
        is_total_row = _is_total_row(row)
        pdf.set_font("Helvetica", "B" if is_total_row else "", 8)
        for col_idx, cell in enumerate(row[: len(headers)]):
            max_widths[col_idx] = max(
                max_widths[col_idx],
                pdf.get_string_width(str(cell)) + cell_padding,
            )

    total = sum(max_widths)
    if total <= 0:
        even_width = table_width / len(headers)
        return tuple(even_width for _ in headers)

    if shrink_to_fit and total <= table_width:
        return tuple(max_widths)

    scale = table_width / total
    return tuple(width * scale for width in max_widths)


def _content_width(pdf: FPDF) -> float:
    return pdf.w - pdf.l_margin - pdf.r_margin


def _remaining_page_height(pdf: FPDF) -> float:
    return pdf.page_break_trigger - pdf.get_y()


def _ensure_space(pdf: FPDF, needed_height: float) -> None:
    """Start a new page if a block would be split across pages."""
    if needed_height > 0 and _remaining_page_height(pdf) < needed_height:
        pdf.add_page()


def _text_block_height(
    pdf: FPDF,
    text: str,
    font_size: int,
    line_height: float,
    font_style: str = "",
) -> float:
    pdf.set_font("Helvetica", font_style, font_size)
    lines = pdf.multi_cell(_content_width(pdf), line_height, text, dry_run=True, output="LINES")
    return len(lines) * line_height


def _title_block_height(pdf: FPDF, title: str) -> float:
    title_height = _text_block_height(pdf, title, 13, 9, "B")
    return title_height + 6.0


def _intro_block_height(pdf: FPDF, intro_lines: list[str]) -> float:
    if not intro_lines:
        return 0.0
    pdf.set_font("Helvetica", "", 10)
    total = 0.0
    for line in intro_lines:
        total += _text_block_height(pdf, line, 10, 7)
    return total + 6.0


def _draw_intro_block(pdf: FPDF, intro_lines: list[str]) -> None:
    if not intro_lines:
        return
    pdf.set_font("Helvetica", "", 10)
    for line in intro_lines:
        pdf.cell(0, 7, line, align="L", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(6)


def _footer_block_height(pdf: FPDF, footer_lines: list[str]) -> float:
    if not footer_lines:
        return 0.0
    pdf.set_font("Helvetica", "", 10)
    total = 5.0
    for line in footer_lines:
        total += _text_block_height(pdf, line, 10, 7)
    return total


def _draw_title_block(pdf: FPDF, title: str) -> None:
    pdf.set_font("Helvetica", "B", 13)
    pdf.multi_cell(_content_width(pdf), 9, title, align="L")
    pdf.ln(6)


def _draw_footer_block(pdf: FPDF, footer_lines: list[str]) -> None:
    if not footer_lines:
        return
    pdf.ln(5)
    pdf.set_font("Helvetica", "", 10)
    for line in footer_lines:
        pdf.cell(0, 7, line, align="L", new_x=XPos.LMARGIN, new_y=YPos.NEXT)


def _truncate_cell(text: str, max_len: int = 28) -> str:
    text = str(text)
    return text if len(text) <= max_len else text[: max_len - 3] + "..."


def _category_break_row_indices(rows: list[list[str]], type_column: int = 1) -> set[int]:
    """Return data row indices that end a category group."""
    if not rows:
        return set()

    breaks: set[int] = set()
    for i in range(len(rows)):
        if i == len(rows) - 1 or rows[i][type_column] != rows[i + 1][type_column]:
            breaks.add(i)
    return breaks


def _parts_table_borders_layout(rows: list[list[str]]):
    """Border layout with a thick line after each part-type group."""
    from fpdf.enums import TableBorderStyle, TableBordersLayout, TableCellStyle

    break_rows = _category_break_row_indices(rows)

    class CategoryBreakBordersLayout(TableBordersLayout):
        def cell_style_getter(
            self,
            row_idx: int,
            col_idx: int,
            col_pos: int,
            num_heading_rows: int,
            num_rows: int,
            num_col_idx: int,
            num_col_pos: int,
        ) -> TableCellStyle:
            style = TableBordersLayout.ALL.cell_style_getter(
                row_idx,
                col_idx,
                col_pos,
                num_heading_rows,
                num_rows,
                num_col_idx,
                num_col_pos,
            )
            data_row_idx = row_idx - int(num_heading_rows)
            if data_row_idx in break_rows:
                return TableCellStyle(
                    left=style.left,
                    bottom=TableBorderStyle(thickness=1.0),
                    right=style.right,
                    top=style.top,
                )
            return style

    return CategoryBreakBordersLayout()


def _is_total_row(row: list[str]) -> bool:
    if not row:
        return False
    label = str(row[0]).strip().lower()
    return label in ("total", "grand total")


def _draw_table(pdf: FPDF, headers: list[str], rows: list[list[str]]) -> None:
    """Draw a table with styled headers and alternating row colors."""
    if not headers:
        return

    from fpdf.fonts import FontFace
    from fpdf.enums import TableBordersLayout

    col_aligns, full_text_columns = _table_column_alignments(headers)
    max_page_width = _content_width(pdf)
    is_parts_table = headers == PARTS_TABLE_HEADERS
    is_modules_style_table = headers in MODULES_STYLE_TABLE_HEADERS

    if is_parts_table:
        col_widths = PARTS_TABLE_COL_WIDTHS
        table_width = max_page_width
    else:
        col_widths = _auto_col_widths(
            pdf,
            headers,
            rows,
            table_width=max_page_width,
            shrink_to_fit=is_modules_style_table,
        )
        table_width = sum(col_widths) if is_modules_style_table else max_page_width
    headings_style = FontFace(
        emphasis="BOLD",
        size_pt=9,
        color=(255, 255, 255),
        fill_color=(31, 78, 121),
    )
    cell_style = FontFace(size_pt=8)
    bold_cell_style = FontFace(emphasis="BOLD", size_pt=8)

    pdf.set_font("Helvetica", "", 8)

    borders_layout = (
        _parts_table_borders_layout(rows)
        if is_parts_table
        else TableBordersLayout.ALL
    )

    with pdf.table(
        width=table_width,
        col_widths=col_widths,
        align="LEFT" if is_modules_style_table else "CENTER",
        line_height=7,
        first_row_as_headings=True,
        headings_style=headings_style,
        cell_fill_color=(235, 241, 250),
        cell_fill_mode="ROWS",
        borders_layout=borders_layout,
    ) as table:
        header_row = table.row()
        for col_idx, header in enumerate(headers):
            header_row.cell(header, align=col_aligns[col_idx])

        for row in rows:
            is_total_row = _is_total_row(row)
            row_style = bold_cell_style if is_total_row else cell_style
            data_row = table.row()
            for col_idx, cell in enumerate(row):
                text = cell if col_idx in full_text_columns else _truncate_cell(cell)
                data_row.cell(text, style=row_style, align=col_aligns[col_idx])


def _render_pdf_section(pdf: FPDF, section: ReportSection, section_index: int) -> None:
    """Render one numbered section: title above the table, footer text below."""
    if section.page_break_before and section_index > 0:
        pdf.add_page()

    _ensure_space(pdf, _title_block_height(pdf, section.title))
    _draw_title_block(pdf, section.title)

    if section.intro_table_headers and section.intro_table_rows:
        _draw_table(pdf, section.intro_table_headers, section.intro_table_rows)
        pdf.ln(6)
    elif section.intro:
        _ensure_space(pdf, _intro_block_height(pdf, section.intro))
        _draw_intro_block(pdf, section.intro)

    if section.headers and section.rows:
        _draw_table(pdf, section.headers, section.rows)
    elif section.headers:
        _draw_table(pdf, section.headers, [])

    if section.footer_table_headers and section.footer_table_rows:
        pdf.ln(6)
        _draw_table(pdf, section.footer_table_headers, section.footer_table_rows)

    if section.footer:
        _ensure_space(pdf, _footer_block_height(pdf, section.footer))
        _draw_footer_block(pdf, section.footer)

    pdf.ln(10)


def _open_pdf(pdf_path: Path) -> None:
    """Open the PDF with the system's default viewer."""
    try:
        if sys.platform == "win32":
            os.startfile(pdf_path)
        elif sys.platform == "darwin":
            subprocess.run(["open", str(pdf_path)], check=False)
        else:
            subprocess.run(["xdg-open", str(pdf_path)], check=False)
    except OSError as exc:
        print(f"Could not open PDF: {exc}")


def save_report_pdf(title: str, sections: list[ReportSection], project_name: str) -> Path:
    """Save the summary report as a PDF in the reports folder."""
    pdf_path = _report_pdf_path(project_name)

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    report_title_height = _text_block_height(pdf, title, 14, 10, "B") + 6
    _ensure_space(pdf, report_title_height)
    pdf.set_font("Helvetica", "B", 14)
    pdf.multi_cell(_content_width(pdf), 10, title, align="C")
    pdf.ln(6)

    for index, section in enumerate(sections):
        _render_pdf_section(pdf, section, index)

    pdf.output(str(pdf_path))
    return pdf_path


def print_table(
    project_name: str,
    module_type_names: list[str],
    part_category: dict[str, str],
    module_counts: dict[str, int],
    total_modules: int,
    total_parts: dict[str, int],
    inventory_stock: dict[str, int],
) -> Path:
    """Save the summary as a PDF report."""
    title, sections = build_summary_report(
        project_name,
        module_type_names,
        part_category,
        module_counts,
        total_modules,
        total_parts,
        inventory_stock,
    )
    pdf_path = save_report_pdf(title, sections, project_name)
    _open_pdf(pdf_path)
    return pdf_path


def main() -> None:
    module_type_names, part_category, module_types_bom = load_bom_from_excel()
    inventory_stock = load_inventory_from_excel()
    project_name = get_project_name()
    module_counts = get_module_counts(module_type_names)

    if sum(module_counts.values()) == 0:
        return

    total_modules, total_parts, _ = calculate(module_counts, module_types_bom)
    print_table(
        project_name,
        module_type_names,
        part_category,
        module_counts,
        total_modules,
        total_parts,
        inventory_stock,
    )


if __name__ == "__main__":
    main()
